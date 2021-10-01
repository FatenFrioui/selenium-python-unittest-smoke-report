import time

from adodbapi.examples.xls_read import sheet
from openpyxl import load_workbook
from selenium import webdriver

driver=webdriver.Chrome(executable_path="C:\webdriver\chromedriver.exe")
# driver=webdriver.Chrome(path+"\Drivers\chromedriver.exe")
driver.get("http://www.uitestingplayground.com/")

driver.find_element_by_xpath("//a[contains(text(),'Sample App')]").click()
driver.maximize_window()
filepath='C:\\fichiers\\testfile.xlsx'
fichier=load_workbook(filepath)
sheet=fichier.active
for i in range(2,10):
    user=sheet.cell(row= i ,column= 1)
    pwd = sheet.cell(row=i, column=2)
    if user.value!=None:
        driver.find_element_by_name("UserName").send_keys("user")
    if pwd.value != None:
        driver.find_element_by_name("Password").send_keys("pwd")

    driver.find_element_by_id("login").click()

    msg = driver.find_element_by_id('loginstatus').text
    bouton=driver.find_element_by_id("login").text
    if bouton=="Log Out":
        driver.find_element_by_id("login").click()
    sheet.cell(row=i,column=4).value=msg
    print(i,user.value,pwd.value,msg)
fichier.save(filepath)
driver.close()


assert msg in 'Welcome, Faten!'



# assert (condition ex 2+2=4),"msg exp the user did not log in"
# assert (msg="welcome, Faten!"),"msg dans le concole"

print('hi')